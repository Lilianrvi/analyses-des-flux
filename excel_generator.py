# excel_generator.py
import os
from openpyxl import load_workbook
import config

def load_template_workbook():
    template_path = "template.xlsx"
    if not os.path.exists(template_path):
        raise FileNotFoundError("Le fichier 'template.xlsx' est introuvable.")
    wb = load_workbook(filename=template_path)
    if config.EXCEL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"La feuille '{config.EXCEL_SHEET_NAME}' est manquante.")
    return wb

def set_cell_value(ws, cell_coord, value):
    try:
        numeric_value = int(value)
    except (ValueError, TypeError):
        numeric_value = None
    for merged_range in ws.merged_cells.ranges:
        if cell_coord in merged_range:
            ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
                value=numeric_value if numeric_value is not None else value
            )
            return
    ws[cell_coord] = numeric_value if numeric_value is not None else value

def fill_excel_workbook(wb, data_par_produit, client_info):
    ws = wb[config.EXCEL_SHEET_NAME]
    # Champs globaux
    set_cell_value(ws, config.GLOBAL_FIELDS["Nom du client"], client_info.get("Nom du client", ""))
    comptes = client_info.get("Comptes clients", [])
    set_cell_value(ws, config.GLOBAL_FIELDS["Comptes clients"], ", ".join(comptes) if comptes else "")
    set_cell_value(ws, config.GLOBAL_FIELDS["Périodicité"], client_info.get("Périodicité", ""))
    
    # Pour I6, extraire le mois (nombre) de la dernière date de la période (format mm/aaaa)
    period_str = client_info.get("Périodicité", "")
    parts = period_str.split()
    if len(parts) < 9:
        raise ValueError("Format de période invalide.")
    quoted_date = parts[8]  # ex: "12/2024"
    try:
        last_month = int(quoted_date.split("/")[0])
    except Exception:
        last_month = 0
    set_cell_value(ws, config.GLOBAL_FIELDS["Dernier mois"], last_month)
    
    # Pour les en-têtes, on utilise la période du premier fichier Excel importé
    # On attend le format "Du mm/aaaa au mm/aaaa et du mm/aaaa au mm/aaaa"
    quoted_date_N_1 = parts[1]  # ex: "01/2023" pour N-1
    quoted_date_N   = parts[8]  # ex: "12/2024" pour N
    year_N_1 = quoted_date_N_1.split("/")[1]
    year_N = quoted_date_N.split("/")[1]
    if year_N_1 == year_N:
        raise ValueError("Les années de comparaison sont identiques dans la période.")
    header_val_N_1 = int(year_N_1)
    header_val_N = int(year_N)
    cells_N   = ["D9", "F9", "L9", "N9", "D36", "F36", "L36", "N36", "R9", "S37", "U37", "W37"]
    cells_N_1 = ["E9", "G9", "M9", "O9", "E36", "G36", "M36", "O36", "T37", "V37", "X37"]
    for cell in cells_N:
        set_cell_value(ws, cell, header_val_N)
    for cell in cells_N_1:
        set_cell_value(ws, cell, header_val_N_1)
    
    # Remplissage des données variables uniquement pour les cellules spécifiées
    for tableau, annees in config.get_excel_structure(date_N_1, date_N):
        for annee, produits in annees.items():
            for produit, cell in produits.items():
                if produit in data_par_produit and annee in data_par_produit[produit]:
                    if tableau == "RC":
                        valeur = data_par_produit[produit][annee].get("RC", 0)
                    elif tableau == "Tonnage":
                        valeur = data_par_produit[produit][annee].get("Tonnage", 0)
                    elif tableau == "CA":
                        valeur = data_par_produit[produit][annee].get("CA", 0)
                    else:
                        valeur = 0
                else:
                    valeur = 0
                set_cell_value(ws, cell, valeur)
    return wb

# Fonction dédiée pour l'addition des fichiers Excel
def fill_excel_workbook_addition(wb, combined_data, period, client_name, client_accounts):
    ws = wb[config.EXCEL_SHEET_NAME]
    # Remplir les champs globaux à partir du premier fichier
    ws[config.GLOBAL_FIELDS["Nom du client"]].value = client_name
    ws[config.GLOBAL_FIELDS["Comptes clients"]].value = client_accounts
    ws[config.GLOBAL_FIELDS["Périodicité"]].value = period
    
    # Reprendre les en-têtes à partir de la période (du premier fichier)
    parts = period.split()
    if len(parts) < 9:
        raise ValueError("Format de période invalide dans le fichier Excel.")
    quoted_date_N_1 = parts[1]  # ex: "01/2023" pour N-1
    quoted_date_N = parts[8]    # ex: "12/2024" pour N
    year_N_1 = quoted_date_N_1.split("/")[1]
    year_N = quoted_date_N.split("/")[1]
    if year_N_1 == year_N:
        raise ValueError("Les années de comparaison sont identiques dans le fichier Excel.")
    header_val_N_1 = int(year_N_1)
    header_val_N = int(year_N)
    cells_N   = ["D9", "F9", "L9", "N9", "D36", "F36", "L36", "N36", "R9", "S37", "U37", "W37"]
    cells_N_1 = ["E9", "G9", "M9", "O9", "E36", "G36", "M36", "O36", "T37", "V37", "X37"]
    for cell in cells_N:
        ws[cell].value = header_val_N
    for cell in cells_N_1:
        ws[cell].value = header_val_N_1
    try:
        header_I6 = int(quoted_date_N.split("/")[0])
    except Exception:
        header_I6 = 0
    ws[config.GLOBAL_FIELDS["Dernier mois"]].value = header_I6
    
    # Remplissage des cellules des tableaux à partir de combined_data
    for table, years in config.get_excel_structure(date_N_1, date_N):
        for year, products in years.items():
            for product, cell in products.items():
                ws[cell].value = combined_data[table][year][product]
    return wb
